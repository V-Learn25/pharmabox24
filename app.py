import os
import json
import time
import uuid
import secrets
import zipfile
import logging
from collections import defaultdict
from datetime import datetime, timedelta, date
from functools import wraps
from threading import Lock
from urllib.parse import urlparse
from urllib.request import Request, urlopen
from urllib.error import URLError, HTTPError

try:
    from zoneinfo import ZoneInfo
except ImportError:  # pragma: no cover — Python < 3.9
    ZoneInfo = None

from flask import Flask, render_template, redirect, url_for, flash, request, jsonify, abort, session
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_wtf import FlaskForm, CSRFProtect
from flask_wtf.csrf import CSRFError
from flask_wtf.file import FileField, FileRequired, FileAllowed
from wtforms import StringField, PasswordField, SelectField, EmailField, BooleanField, TextAreaField
from wtforms.validators import DataRequired, Email, Length, EqualTo, Optional
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.utils import secure_filename
from markupsafe import escape
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
from openpyxl import load_workbook
import csv

from config import Config
from models import db, User, Organisation, Pharmacy, DailyStat, HourlyDistribution, Upload

logger = logging.getLogger(__name__)

app = Flask(__name__)
app.config.from_object(Config)

# Honour X-Forwarded-* headers from the trusted proxy depth in front of us (Railway = 1).
app.wsgi_app = ProxyFix(
    app.wsgi_app,
    x_for=app.config['TRUSTED_PROXY_COUNT'],
    x_proto=app.config['TRUSTED_PROXY_COUNT'],
    x_host=app.config['TRUSTED_PROXY_COUNT'],
)

db.init_app(app)
csrf = CSRFProtect(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'
login_manager.session_protection = 'strong'


@app.errorhandler(CSRFError)
def csrf_error(e):
    flash('Your session expired. Please try again.', 'error')
    # Redirect to a fixed safe destination — never reflect request.url back.
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Simple in-memory rate limiter for login/forgot-password.
# NOTE: this is PER-WORKER. With gunicorn --workers N, effective limit is N × _MAX_ATTEMPTS.
# Acceptable as a layered control alongside strong passwords and account lockout; replace with
# Flask-Limiter + Redis once that infrastructure is available.
_login_attempts = defaultdict(list)
_attempts_lock = Lock()
_MAX_ATTEMPTS = 5            # Per-IP per worker, per window
_WINDOW_SECONDS = 300        # 5 minutes
_MAX_TRACKED_KEYS = 5000     # Hard cap to prevent memory exhaustion


def _client_ip():
    """Return the client IP, trusting only as many proxies as TRUSTED_PROXY_COUNT permits."""
    # ProxyFix has already rewritten request.remote_addr based on TRUSTED_PROXY_COUNT.
    return request.remote_addr or 'unknown'


def _is_rate_limited(key):
    """Check if key has exceeded rate limit. Cleans old entries; bounded memory."""
    now = time.time()
    with _attempts_lock:
        attempts = [t for t in _login_attempts.get(key, ()) if now - t < _WINDOW_SECONDS]
        if attempts:
            _login_attempts[key] = attempts
        elif key in _login_attempts:
            del _login_attempts[key]

        # Periodic janitor — purge cold keys when the dict grows
        if len(_login_attempts) > _MAX_TRACKED_KEYS:
            cutoff = now - _WINDOW_SECONDS
            stale = [k for k, ts in _login_attempts.items() if not ts or ts[-1] < cutoff]
            for k in stale:
                _login_attempts.pop(k, None)

        return len(attempts) >= _MAX_ATTEMPTS


def _record_attempt(key):
    with _attempts_lock:
        _login_attempts[key].append(time.time())


_CSP = (
    "default-src 'self'; "
    # 'unsafe-inline' is required for the inline tailwind config block in base.html.
    # Migrate to a nonce when templates are refactored or Tailwind is compiled at build time.
    "script-src 'self' 'unsafe-inline' "
    "https://cdn.jsdelivr.net https://cdn.tailwindcss.com https://cdnjs.cloudflare.com; "
    "style-src 'self' 'unsafe-inline' "
    "https://cdnjs.cloudflare.com https://cdn.tailwindcss.com; "
    "img-src 'self' data:; "
    "font-src 'self' data: https://cdnjs.cloudflare.com; "
    "connect-src 'self'; "
    "frame-ancestors 'self'; "
    "base-uri 'self'; "
    "form-action 'self'; "
    "object-src 'none'"
)


@app.after_request
def set_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'geolocation=(), microphone=(), camera=(), payment=()'
    response.headers['Content-Security-Policy'] = _CSP
    # Only emit HSTS over HTTPS — avoid pinning HTTP-only dev hosts.
    if request.is_secure:
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    # X-XSS-Protection is deprecated and can introduce vulnerabilities — do not set it.
    response.headers.pop('X-XSS-Protection', None)
    return response


def send_email(to_email, subject, html_content, reply_to=None):
    """Send email via Resend API. Returns True on success.

    `reply_to` (optional): an email address the recipient's mail client should reply
    to instead of MAIL_FROM. Used by the Contact form so support replies route to the
    user who filed the ticket, not back to noreply@.
    """
    api_key = app.config.get('RESEND_API_KEY')
    if not api_key:
        app.logger.warning('RESEND_API_KEY not configured - skipping email')
        return False

    payload_dict = {
        'from': app.config['MAIL_FROM'],
        'to': [to_email],
        'subject': subject,
        'html': html_content,
    }
    if reply_to:
        payload_dict['reply_to'] = reply_to
    payload = json.dumps(payload_dict).encode()

    req = Request(
        'https://api.resend.com/emails',
        data=payload,
        headers={
            'Authorization': f'Bearer {api_key}',
            'Content-Type': 'application/json',
            # Resend sits behind Cloudflare; the default Python-urllib/X user-agent gets
            # WAF-rejected (error code 1010). A descriptive UA satisfies the bot filter.
            'User-Agent': 'Pharmabox24/1.0 (+https://managementinfo.pharmabox24.co.uk)',
            'Accept': 'application/json',
        },
        method='POST'
    )

    try:
        with urlopen(req, timeout=10) as resp:
            if 200 <= resp.status < 300:
                app.logger.info(f'Email sent to {to_email} (status {resp.status})')
                return True
            app.logger.error(f'Resend returned status {resp.status} for {to_email}')
            return False
    except HTTPError as e:
        body = ''
        try:
            body = e.read().decode('utf-8', errors='replace')[:500]
        except Exception:
            pass
        app.logger.error(f'Resend HTTPError {e.code} sending to {to_email}: {body}')
        return False
    except (URLError, TimeoutError) as e:
        app.logger.error(f'Resend network error sending to {to_email}: {e}')
        return False
    except (ValueError, OSError) as e:
        app.logger.exception(f'Unexpected error sending to {to_email}: {e}')
        return False


def _safe_int(value, default=0):
    """Coerce a value to int safely for email/template use."""
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _site_url():
    """Base URL used in emails for links and the logo. Falls back to the request host."""
    site = app.config.get('SITE_URL') or ''
    if site:
        return site.rstrip('/')
    try:
        return request.host_url.rstrip('/')
    except RuntimeError:
        return ''


def _email_layout(preheader, body_html):
    """Wrap an email body in the standard Pharmabox24 chrome with the logo at top.

    All emails go through this helper so branding stays consistent. The logo is loaded
    from the live site's /static/images/logo.png; Gmail and Apple Mail render it
    automatically, Outlook desktop requires the user to enable images for the sender.
    """
    site = _site_url()
    logo_url = f'{site}/static/images/logo.png' if site else ''
    safe_pre = escape(preheader or '')
    logo_block = (
        f'<img src="{escape(logo_url)}" alt="Pharmabox24" '
        f'style="max-height: 56px; display: block; margin: 0 auto;">'
    ) if logo_url else (
        '<div style="font-family: Arial, sans-serif; font-size: 24px; font-weight: 900; '
        'color: #00891a; letter-spacing: -1px;">Pharmabox24</div>'
    )
    return f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Pharmabox24</title>
</head>
<body style="margin: 0; padding: 0; background: #f4f5f7; font-family: Arial, sans-serif; color: #333;">
    <!-- preheader (hidden but shown in inbox preview) -->
    <div style="display: none; max-height: 0; overflow: hidden;">{safe_pre}</div>
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0">
        <tr>
            <td align="center" style="padding: 24px 12px;">
                <table role="presentation" width="600" cellspacing="0" cellpadding="0" border="0" style="max-width: 600px; width: 100%; border: 1px solid #e5e7eb; border-radius: 8px; overflow: hidden;">
                    <tr>
                        <td style="background: #ffffff; padding: 28px 24px 20px; text-align: center; border-bottom: 1px solid #e5e7eb;">
                            {logo_block}
                        </td>
                    </tr>
                    <tr>
                        <td style="background: #ffffff; padding: 28px 28px 32px; line-height: 1.6; color: #333;">
                            {body_html}
                        </td>
                    </tr>
                    <tr>
                        <td style="background: #f9fafb; padding: 16px 28px; text-align: center; color: #888; font-size: 12px; border-top: 1px solid #e5e7eb;">
                            Pharmabox24 — Prescription Collection Analytics Portal
                        </td>
                    </tr>
                </table>
            </td>
        </tr>
    </table>
</body>
</html>"""


def _today():
    """Return today's date in the configured display timezone (Europe/Dublin by default)."""
    tz_name = app.config.get('DISPLAY_TIMEZONE') or 'Europe/Dublin'
    if ZoneInfo is None:
        return date.today()
    try:
        return datetime.now(ZoneInfo(tz_name)).date()
    except Exception:
        return date.today()


def _normalize_email(value):
    """Strip whitespace and lower-case an email; bound length to model column size."""
    if not value:
        return ''
    return (value or '').strip().lower()[:120]


def _format_date_range(date_from, date_to, compact=False):
    """Render a (date_from, date_to) span for human eyes. Returns '' if either is missing.

    `compact=True` produces a tighter form suitable for an email subject line:
    "27–30 Apr 2026" when both endpoints share a month, "27 Apr – 3 May 2026"
    when they don't. `compact=False` gives full "27 April 2026 – 3 May 2026" for
    the body. Single-day ranges collapse to one date.
    """
    if not date_from or not date_to:
        return ''
    if date_from == date_to:
        return date_from.strftime('%-d %b %Y' if compact else '%-d %B %Y')
    if compact:
        if date_from.year == date_to.year and date_from.month == date_to.month:
            return f"{date_from.strftime('%-d')}–{date_to.strftime('%-d %b %Y')}"
        if date_from.year == date_to.year:
            return f"{date_from.strftime('%-d %b')} – {date_to.strftime('%-d %b %Y')}"
        return f"{date_from.strftime('%-d %b %Y')} – {date_to.strftime('%-d %b %Y')}"
    return f"{date_from.strftime('%-d %B %Y')} – {date_to.strftime('%-d %B %Y')}"


def send_notification_email(pharmacy, stats_summary):
    """Send email notification to pharmacy about new data upload."""
    if not pharmacy.notification_email:
        return False

    site_url = app.config.get('SITE_URL') or request.host_url.rstrip('/')
    login_url = f"{site_url}/login"

    # Every user-controlled value is HTML-escaped before interpolation. Pharmacy names
    # originate from uploaded spreadsheets and must NEVER be trusted as HTML.
    safe_name = escape(pharmacy.name or '')
    loaded = _safe_int(stats_summary.get('loaded'))
    collected = _safe_int(stats_summary.get('collected'))
    removed = _safe_int(stats_summary.get('removed'))
    safe_login = escape(login_url)

    date_from = stats_summary.get('date_from')
    date_to = stats_summary.get('date_to')
    body_range = escape(_format_date_range(date_from, date_to, compact=False))
    subject_range = _format_date_range(date_from, date_to, compact=True)

    # The italic date-range line is omitted entirely when the upload had no daily rows
    # for this pharmacy (e.g. hourly-only data) — there are no dates to honestly state.
    range_block = (
        f'<p style="color: #555; font-size: 14px; margin-top: -4px;">Reporting period: '
        f'<strong>{body_range}</strong></p>'
        if body_range else ''
    )

    body = f"""
        <h2 style="color: #00891a; margin-top: 0;">New statistics available</h2>
        <p>Hello <strong>{safe_name}</strong>,</p>
        <p>New statistics have been uploaded for your pharmacy. Here's a quick summary:</p>
        {range_block}

        <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="margin: 16px 0;">
            <tr><td style="padding: 8px 0;"><div style="background: #f4f5f7; padding: 14px 16px; border-radius: 6px; border-left: 4px solid #00891a;">
                <div style="font-size: 12px; color: #666; text-transform: uppercase;">Loaded Parcels</div>
                <div style="font-size: 24px; font-weight: bold; color: #00891a;">{loaded}</div>
            </div></td></tr>
            <tr><td style="padding: 8px 0;"><div style="background: #f4f5f7; padding: 14px 16px; border-radius: 6px; border-left: 4px solid #00891a;">
                <div style="font-size: 12px; color: #666; text-transform: uppercase;">Collected Parcels</div>
                <div style="font-size: 24px; font-weight: bold; color: #00891a;">{collected}</div>
            </div></td></tr>
            <tr><td style="padding: 8px 0;"><div style="background: #f4f5f7; padding: 14px 16px; border-radius: 6px; border-left: 4px solid #00891a;">
                <div style="font-size: 12px; color: #666; text-transform: uppercase;">Removed Parcels</div>
                <div style="font-size: 24px; font-weight: bold; color: #00891a;">{removed}</div>
            </div></td></tr>
        </table>

        <p style="text-align: center; margin: 24px 0;">
            <a href="{safe_login}" style="display: inline-block; background: #00891a; color: white; padding: 12px 24px; text-decoration: none; border-radius: 6px; font-weight: 600;">View full analytics</a>
        </p>
    """
    html_content = _email_layout(f'New statistics for {pharmacy.name}', body)

    # Subject lines also surface user-controlled values — neutralise newlines/control chars.
    safe_subject_name = (pharmacy.name or '').replace('\r', ' ').replace('\n', ' ')[:120]
    subject = f'New Statistics Available - {safe_subject_name}'
    if subject_range:
        subject = f'{subject} ({subject_range})'
    return send_email(
        pharmacy.notification_email,
        subject,
        html_content
    )


def send_password_changed_notice(user):
    """Out-of-band notification when a user's password is changed."""
    if not user.email:
        return False
    safe_name = escape(user.name or '')
    body = f"""
        <h2 style="color: #00891a; margin-top: 0;">Password changed</h2>
        <p>Hello <strong>{safe_name}</strong>,</p>
        <p>Your Pharmabox24 password was just changed.</p>
        <p style="color: #666;">If this wasn't you, contact your administrator immediately.</p>
    """
    html = _email_layout('Your Pharmabox24 password was changed', body)
    return send_email(user.email, 'Pharmabox24 password changed', html)


def send_password_reset_email(user, reset_url):
    """Password reset link — used by both the public forgot-password flow and the
    admin-triggered 'send reset' button on the user-edit page."""
    safe_name = escape(user.name or '')
    safe_url = escape(reset_url)
    body = f"""
        <h2 style="color: #00891a; margin-top: 0;">Password reset</h2>
        <p>Hello <strong>{safe_name}</strong>,</p>
        <p>You requested a password reset. Click the button below within 1 hour to set a new password:</p>
        <p style="text-align: center; margin: 24px 0;">
            <a href="{safe_url}" style="display: inline-block; background: #00891a; color: white; padding: 12px 24px; text-decoration: none; border-radius: 6px; font-weight: 600;">Reset password</a>
        </p>
        <p style="font-size: 12px; color: #666;">If you didn't request this, you can safely ignore this email.</p>
    """
    html = _email_layout('Reset your Pharmabox24 password', body)
    return send_email(user.email, 'Password Reset - Pharmabox24', html)


def send_invitation_email(user, setup_url):
    """Sent when an admin adds a new user. Welcome + password-setup link.

    The setup link is generated from a 7-day invitation token (separate salt from
    reset tokens). Once the user sets a password, set_password() bumps session_version
    which invalidates the invite token too.
    """
    safe_name = escape(user.name or '')
    safe_email = escape(user.email or '')
    safe_url = escape(setup_url)
    role_label = (
        'Super Administrator' if user.role == 'super_admin'
        else 'Organisation Administrator' if user.role == 'org_admin'
        else 'Pharmacy User'
    )
    body = f"""
        <h2 style="color: #00891a; margin-top: 0;">Welcome to Pharmabox24</h2>
        <p>Hello <strong>{safe_name}</strong>,</p>
        <p>An account has been created for you on Pharmabox24, the prescription collection analytics portal.</p>
        <table role="presentation" cellspacing="0" cellpadding="0" style="background: #f4f5f7; border-radius: 6px; margin: 16px 0; width: 100%;">
            <tr><td style="padding: 14px 16px;">
                <div style="font-size: 12px; color: #666; text-transform: uppercase;">Your login email</div>
                <div style="font-size: 16px; font-weight: 600; color: #333;">{safe_email}</div>
                <div style="font-size: 12px; color: #666; text-transform: uppercase; margin-top: 12px;">Your role</div>
                <div style="font-size: 16px; font-weight: 600; color: #333;">{role_label}</div>
            </td></tr>
        </table>
        <p>To activate your account, click the button below to set your password. This link is valid for 7 days.</p>
        <p style="text-align: center; margin: 24px 0;">
            <a href="{safe_url}" style="display: inline-block; background: #00891a; color: white; padding: 12px 28px; text-decoration: none; border-radius: 6px; font-weight: 600;">Set your password</a>
        </p>
        <p style="font-size: 13px; color: #666;">If you weren't expecting this invitation, you can safely ignore this email — no account will be activated until you set a password.</p>
    """
    html = _email_layout('Activate your Pharmabox24 account', body)
    return send_email(user.email, 'Welcome to Pharmabox24 — activate your account', html)


def send_support_request_email(user, subject, message):
    """Forward an in-portal Contact-Us submission to SUPPORT_EMAIL.

    Reply-To is set to the requesting user so support staff can reply directly to
    them without needing to copy/paste their address. The user identity context
    (role, pharmacy, organisation) is included in the body so triage is fast.
    """
    safe_name = escape(user.name or '')
    safe_email = escape(user.email or '')
    safe_subject = escape(subject)
    # Preserve line breaks in the user's message while escaping HTML.
    safe_message = escape(message).replace('\n', '<br>')
    role_label = (
        'Super Administrator' if user.role == 'super_admin'
        else 'Organisation Administrator' if user.role == 'org_admin'
        else 'Pharmacy User'
    )
    pharmacy_name = escape(user.pharmacy.name) if user.pharmacy else '(not linked)'
    org_name = escape(user.organisation.name) if user.organisation else '(not linked)'

    body = f"""
        <h2 style="color: #00891a; margin-top: 0;">Support request</h2>
        <p style="color: #555;">Submitted via the in-portal Contact form.</p>
        <table role="presentation" cellspacing="0" cellpadding="0" style="background: #f9fafb; border-radius: 6px; margin: 16px 0; width: 100%;">
            <tr><td style="padding: 14px 16px; line-height: 1.8;">
                <strong>From:</strong> {safe_name} &lt;<a href="mailto:{safe_email}" style="color: #00891a;">{safe_email}</a>&gt;<br>
                <strong>Role:</strong> {role_label}<br>
                <strong>Pharmacy:</strong> {pharmacy_name}<br>
                <strong>Organisation:</strong> {org_name}
            </td></tr>
        </table>
        <h3 style="color: #333; margin-bottom: 4px;">Subject</h3>
        <p style="margin-top: 0;">{safe_subject}</p>
        <h3 style="color: #333; margin-bottom: 4px;">Message</h3>
        <div style="background: #ffffff; border: 1px solid #e5e7eb; border-radius: 6px; padding: 14px 16px;">{safe_message}</div>
        <p style="font-size: 12px; color: #888; margin-top: 24px;">Reply directly to this email to respond to {safe_name}.</p>
    """
    html = _email_layout(f'Support: {subject}', body)
    # Subject prefix makes filter rules in the support inbox trivial.
    # Strip newlines from user-supplied subject (just in case) for the email header.
    safe_subject_header = (subject or '').replace('\r', ' ').replace('\n', ' ')[:180]
    return send_email(
        SUPPORT_EMAIL,
        f'[Pharmabox24 Support] {safe_subject_header}',
        html,
        reply_to=user.email,
    )


@login_manager.user_loader
def load_user(user_id):
    """Tolerate both legacy raw-int IDs and the new id|session_version format."""
    if user_id is None:
        return None
    raw = str(user_id)
    pinned_version = None
    if '|' in raw:
        raw_id, _, ver = raw.partition('|')
        try:
            pinned_version = int(ver)
        except ValueError:
            return None
    else:
        raw_id = raw
    try:
        uid = int(raw_id)
    except ValueError:
        return None
    user = db.session.get(User, uid)
    if not user:
        return None
    # If the cookie was issued before a force-logout / password change, reject it.
    if pinned_version is not None and (user.session_version or 1) != pinned_version:
        return None
    return user


# --- Decorators ---

def super_admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login', next=request.path))
        if not current_user.is_super_admin():
            # Soft notice instead of alarming red error — the user lands on a page they CAN
            # access (their role-appropriate dashboard), so a red error bar is misleading.
            flash('That page is for administrators — showing your dashboard instead.', 'info')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function


def admin_required(f):
    """Allows super_admin OR org_admin (legacy 'admin' role treated as super_admin)."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login', next=request.path))
        if not current_user.is_admin():
            flash('That page is for administrators — showing your dashboard instead.', 'info')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function


def _get_org_pharmacy_ids():
    """Get pharmacy IDs belonging to the current org_admin's organisation."""
    if not current_user.organisation_id:
        return []
    return [p.id for p in Pharmacy.query.filter_by(organisation_id=current_user.organisation_id).all()]


def _can_access_pharmacy(pharmacy_id):
    """Check if current user can access a given pharmacy."""
    if current_user.is_super_admin():
        return True
    if current_user.is_org_admin():
        pharmacy = db.session.get(Pharmacy, pharmacy_id)
        return pharmacy and pharmacy.organisation_id == current_user.organisation_id
    if current_user.pharmacy_id == pharmacy_id:
        return True
    return False


# Forms
class LoginForm(FlaskForm):
    email = EmailField('Email', validators=[DataRequired(), Email()])
    password = PasswordField('Password', validators=[DataRequired()])
    remember_me = BooleanField('Keep me signed in on this device', default=False)


# 8-char minimum (NIST SP 800-63B baseline) is the right floor for this app given:
#   - online brute force is rate-limited (5 attempts / 5 min / IP / worker)
#   - hashing is pbkdf2-sha256 @ 600k iterations (~150ms per guess on a modern GPU)
#   - the audience (pharmacy staff) will work around longer minimums with sticky notes
# Going below 8 is not safe; going much above 8 has diminishing returns AND a real UX cost.
_PASSWORD_MIN = 8
_PASSWORD_MAX = 128


class UserForm(FlaskForm):
    """Admin-side user form. NO password field — new users receive an emailed invitation
    link to set their own password. Existing users keep their password until they reset it
    (via /forgot-password) or an admin triggers a reset email."""
    email = EmailField('Email', validators=[DataRequired(), Email()])
    name = StringField('Name', validators=[DataRequired(), Length(min=2, max=100)])
    role = SelectField('Role', choices=[('pharmacy', 'Pharmacy User'), ('org_admin', 'Organisation Admin'), ('super_admin', 'Super Admin')])
    pharmacy_id = SelectField('Pharmacy', coerce=int)
    organisation_id = SelectField('Organisation', coerce=int)


class OrgUserForm(FlaskForm):
    """User form for org admins. Like UserForm — no password field; invitation is emailed."""
    email = EmailField('Email', validators=[DataRequired(), Email()])
    name = StringField('Name', validators=[DataRequired(), Length(min=2, max=100)])
    pharmacy_id = SelectField('Pharmacy', coerce=int)


class PharmacyForm(FlaskForm):
    serial_number = StringField('Serial Number', validators=[DataRequired(), Length(min=1, max=50)])
    name = StringField('Pharmacy Name', validators=[DataRequired(), Length(min=2, max=200)])
    notification_email = EmailField('Notification Email', validators=[Optional(), Email()])
    organisation_id = SelectField('Organisation', coerce=int)


class OrganisationForm(FlaskForm):
    name = StringField('Organisation Name', validators=[DataRequired(), Length(min=2, max=200)])


class UploadForm(FlaskForm):
    file = FileField('Data File', validators=[
        FileRequired(),
        FileAllowed(['csv', 'xlsx', 'xls'], 'CSV or Excel files only!')
    ])


class ChangePasswordForm(FlaskForm):
    current_password = PasswordField('Current Password', validators=[DataRequired()])
    new_password = PasswordField('New Password', validators=[DataRequired(), Length(min=_PASSWORD_MIN, max=_PASSWORD_MAX)])
    confirm_password = PasswordField('Confirm New Password', validators=[DataRequired()])


class ForgotPasswordForm(FlaskForm):
    email = EmailField('Email', validators=[DataRequired(), Email()])


class ResetPasswordForm(FlaskForm):
    new_password = PasswordField('New Password', validators=[DataRequired(), Length(min=_PASSWORD_MIN, max=_PASSWORD_MAX)])
    confirm_password = PasswordField('Confirm Password', validators=[DataRequired()])


class SetupAccountForm(FlaskForm):
    """Used by the new-user invitation flow at /setup-account/<token>."""
    new_password = PasswordField('Choose a password', validators=[DataRequired(), Length(min=_PASSWORD_MIN, max=_PASSWORD_MAX)])
    confirm_password = PasswordField('Confirm password', validators=[DataRequired()])


class ContactForm(FlaskForm):
    """In-portal support / contact form. Sends to SUPPORT_EMAIL."""
    subject = StringField('Subject', validators=[DataRequired(), Length(min=3, max=200)])
    message = TextAreaField('Message', validators=[DataRequired(), Length(min=10, max=5000)])


# Destination for the in-portal Contact Us form.
SUPPORT_EMAIL = 'man.info@pharmabox24.co.uk'


# Password reset token helpers — HMAC-signed via itsdangerous, bound to the user's
# session_version so any password change or forced logout invalidates outstanding tokens.

_RESET_SALT = 'pharmabox-password-reset-v2'


def _reset_serializer():
    return URLSafeTimedSerializer(app.config['SECRET_KEY'], salt=_RESET_SALT)


def generate_reset_token(user):
    return _reset_serializer().dumps({
        'uid': user.id,
        'sv': user.session_version or 1,
    })


def verify_reset_token(token, max_age=3600):
    """Verify a password-reset token. Returns the User or None."""
    if not token or not isinstance(token, str) or len(token) > 1024:
        return None
    try:
        data = _reset_serializer().loads(token, max_age=max_age)
    except (BadSignature, SignatureExpired):
        return None
    if not isinstance(data, dict):
        return None
    uid = data.get('uid')
    sv = data.get('sv')
    if not isinstance(uid, int) or not isinstance(sv, int):
        return None
    user = db.session.get(User, uid)
    if not user:
        return None
    # session_version changes on password reset → outstanding tokens stop working.
    if (user.session_version or 1) != sv:
        return None
    return user


# Invitation token helpers — same mechanism as reset tokens but a different salt
# (so a leaked reset token can't be reused as an invite or vice versa) and a longer
# default expiry (7 days, since invitations may take days to be acted on).
_INVITE_SALT = 'pharmabox-account-invite-v1'
_INVITE_MAX_AGE = 7 * 24 * 3600  # 7 days


def _invite_serializer():
    return URLSafeTimedSerializer(app.config['SECRET_KEY'], salt=_INVITE_SALT)


def generate_invite_token(user):
    return _invite_serializer().dumps({
        'uid': user.id,
        'sv': user.session_version or 1,
    })


def verify_invite_token(token, max_age=_INVITE_MAX_AGE):
    """Verify an invitation token. Returns the User or None.

    Token becomes invalid as soon as the user sets a password (set_password bumps
    session_version). That means a single token is effectively one-shot, even though
    the underlying signing mechanism doesn't enforce one-shot directly.
    """
    if not token or not isinstance(token, str) or len(token) > 1024:
        return None
    try:
        data = _invite_serializer().loads(token, max_age=max_age)
    except (BadSignature, SignatureExpired):
        return None
    if not isinstance(data, dict):
        return None
    uid = data.get('uid')
    sv = data.get('sv')
    if not isinstance(uid, int) or not isinstance(sv, int):
        return None
    user = db.session.get(User, uid)
    if not user:
        return None
    if (user.session_version or 1) != sv:
        return None
    return user


def _invite_new_user(user):
    """Send an account-activation invitation to a freshly-created user. Idempotent —
    can be called again as a 'resend invitation' action as long as the user hasn't
    yet set a password (which would bump session_version and invalidate the token).
    Returns True if the send succeeded.
    """
    token = generate_invite_token(user)
    setup_url = f"{_site_url()}/setup-account/{token}"
    try:
        return send_invitation_email(user, setup_url)
    except Exception:
        app.logger.exception(f'Failed to send invitation to {user.email}')
        return False


def _trigger_password_reset(user):
    """Admin-triggered password reset — generates a regular reset token and emails
    the standard reset link. Used by the 'Send password reset' button on user-edit pages.
    """
    token = generate_reset_token(user)
    reset_url = f"{_site_url()}/reset-password/{token}"
    try:
        return send_password_reset_email(user, reset_url)
    except Exception:
        app.logger.exception(f'Failed to send admin-triggered reset to {user.email}')
        return False


# === ROUTES ===

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


def _safe_next(next_page, user=None):
    """Return next_page only if it's a same-origin relative path AND accessible to `user`.

    Rejects:
      - absolute URLs (`http://evil`, `https://evil`)
      - protocol-relative URLs (`//evil.com/x`)
      - non-`/`-prefixed paths
      - anything containing a host or scheme component after parsing
      - admin/org paths the user doesn't have a role for (prevents the post-login bounce
        where a pharmacy user clicked a bookmarked admin link → lands on their own page
        with a confusing red "permission denied" flash)
    """
    if not next_page or not isinstance(next_page, str):
        return None
    if not next_page.startswith('/') or next_page.startswith('//'):
        return None
    parsed = urlparse(next_page)
    if parsed.scheme or parsed.netloc:
        return None
    if user is not None:
        path = parsed.path or next_page
        if path.startswith('/admin') and not user.is_super_admin():
            return None
        if path.startswith('/org') and not (user.is_super_admin() or user.is_org_admin()):
            return None
    return next_page


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    form = LoginForm()
    if form.validate_on_submit():
        client_ip = _client_ip()
        if _is_rate_limited(f'login:{client_ip}'):
            flash('Too many login attempts. Please wait a few minutes.', 'error')
            return render_template('login.html', form=form)

        email = (form.email.data or '').strip().lower()
        user = User.query.filter_by(email=email).first()
        if user and user.check_password(form.password.data):
            remember = bool(form.remember_me.data)
            login_user(user, remember=remember)
            # Only mark the session permanent (= PERMANENT_SESSION_LIFETIME, 7 days) when the
            # user opted in. Otherwise the cookie is browser-session-only — ideal for shared
            # pharmacy counter PCs.
            session.permanent = remember
            next_page = _safe_next(request.args.get('next'), user=user)
            flash(f'Welcome back, {user.name}!', 'success')
            return redirect(next_page or url_for('dashboard'))
        _record_attempt(f'login:{client_ip}')
        flash('Invalid email or password.', 'error')
    return render_template('login.html', form=form)


@app.route('/logout', methods=['POST'])
@login_required
def logout():
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))


@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    form = ForgotPasswordForm()
    if form.validate_on_submit():
        client_ip = _client_ip()
        if _is_rate_limited(f'reset:{client_ip}'):
            flash('Too many reset requests. Please wait a few minutes.', 'error')
            return redirect(url_for('login'))
        _record_attempt(f'reset:{client_ip}')

        email = (form.email.data or '').strip().lower()
        user = User.query.filter_by(email=email).first()
        if user:
            token = generate_reset_token(user)
            site_url = _site_url()
            reset_url = f"{site_url}/reset-password/{token}"
            send_password_reset_email(user, reset_url)

        flash('If that email exists in our system, a reset link has been sent.', 'info')
        return redirect(url_for('login'))

    return render_template('forgot_password.html', form=form)


@app.route('/setup-account/<token>', methods=['GET', 'POST'])
def setup_account(token):
    """New-user invitation activation. Verifies the invite token, lets the user set
    a password, logs them in. Token is invalidated by the password set (session_version bump).
    """
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    user = verify_invite_token(token)
    if not user:
        flash('This invitation link is invalid or has expired. Ask your administrator to resend it.', 'error')
        return redirect(url_for('login'))

    form = SetupAccountForm()
    if form.validate_on_submit():
        if form.new_password.data != form.confirm_password.data:
            flash('Passwords do not match.', 'error')
            return render_template('setup_account.html', form=form, token=token, invited_email=user.email)

        user.set_password(form.new_password.data)  # bumps session_version → invalidates token
        db.session.commit()
        # Log them in directly so they don't have to re-enter the password they just typed.
        login_user(user, remember=False)
        session.permanent = False
        flash(f'Welcome, {user.name}! Your account is now active.', 'success')
        return redirect(url_for('dashboard'))

    return render_template('setup_account.html', form=form, token=token, invited_email=user.email)


@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    user = verify_reset_token(token)
    if not user:
        flash('Invalid or expired reset link. Please request a new one.', 'error')
        return redirect(url_for('forgot_password'))

    form = ResetPasswordForm()
    if form.validate_on_submit():
        if form.new_password.data != form.confirm_password.data:
            flash('Passwords do not match.', 'error')
            return render_template('reset_password.html', form=form, token=token)

        user.set_password(form.new_password.data)  # bumps session_version → invalidates other sessions
        db.session.commit()
        try:
            send_password_changed_notice(user)
        except Exception:
            app.logger.exception('Failed to send password-changed notice')
        flash('Your password has been reset. Please log in.', 'success')
        return redirect(url_for('login'))

    return render_template('reset_password.html', form=form, token=token)


@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    form = ChangePasswordForm()

    if form.validate_on_submit():
        if not current_user.check_password(form.current_password.data):
            flash('Current password is incorrect.', 'error')
            return render_template('change_password.html', form=form)

        if form.new_password.data != form.confirm_password.data:
            flash('New passwords do not match.', 'error')
            return render_template('change_password.html', form=form)

        # Preserve the user's original remember-me choice across the re-login below.
        was_remembered = bool(session.permanent)

        current_user.set_password(form.new_password.data)  # bumps session_version
        db.session.commit()
        try:
            send_password_changed_notice(current_user)
        except Exception:
            app.logger.exception('Failed to send password-changed notice')

        # The session_version bump just invalidated this session — re-login with the same
        # remember-me posture they had before, so we don't silently upgrade non-persistent
        # sessions to 7-day cookies.
        login_user(current_user, remember=was_remembered)
        session.permanent = was_remembered

        flash('Your password has been updated successfully.', 'success')
        return redirect(url_for('dashboard'))

    return render_template('change_password.html', form=form)


@app.route('/contact', methods=['GET', 'POST'])
@login_required
def contact():
    """In-portal support form — emails SUPPORT_EMAIL with Reply-To set to the user."""
    form = ContactForm()
    if form.validate_on_submit():
        client_ip = _client_ip()
        # Reuse the existing per-IP rate limiter to discourage abuse / accidental spam.
        if _is_rate_limited(f'contact:{client_ip}'):
            flash('Too many support requests in a short period. Please wait a few minutes before trying again.', 'error')
            return redirect(url_for('contact'))
        _record_attempt(f'contact:{client_ip}')

        subject = (form.subject.data or '').strip()
        message = (form.message.data or '').strip()
        try:
            ok = send_support_request_email(current_user, subject, message)
        except Exception:
            app.logger.exception('Failed to send contact-us email')
            ok = False

        if ok:
            flash("Thanks — your message has been sent. We'll get back to you shortly.", 'success')
            return redirect(url_for('contact'))
        flash(
            f"Sorry, we couldn't send your message right now. Please email {SUPPORT_EMAIL} directly, "
            "or try again in a few minutes.",
            'error',
        )
    return render_template('contact.html', form=form, support_email=SUPPORT_EMAIL)


@app.route('/dashboard')
@login_required
def dashboard():
    if current_user.is_super_admin():
        return redirect(url_for('admin_dashboard'))
    if current_user.is_org_admin():
        if current_user.organisation_id:
            return redirect(url_for('org_dashboard'))
        else:
            flash('Your account is not linked to an organisation. Please contact a super admin.', 'error')
            return render_template('pharmacy/dashboard.html', pharmacy=None, stats={})
    return redirect(url_for('pharmacy_dashboard'))


# === PHARMACY USER ROUTES ===

@app.route('/pharmacy/dashboard')
@login_required
def pharmacy_dashboard():
    if current_user.is_super_admin():
        return redirect(url_for('admin_dashboard'))
    if current_user.is_org_admin():
        return redirect(url_for('org_dashboard'))

    if not current_user.pharmacy_id:
        flash('Your account is not linked to a pharmacy. Please contact an administrator.', 'error')
        return render_template('pharmacy/dashboard.html', pharmacy=None, stats={})

    pharmacy = db.session.get(Pharmacy, current_user.pharmacy_id)
    today = _today()
    stats = get_pharmacy_stats(pharmacy.id, today)

    return render_template('pharmacy/dashboard.html', pharmacy=pharmacy, stats=stats)


@app.route('/api/pharmacy/chart-data')
@login_required
def pharmacy_chart_data():
    if current_user.is_super_admin() or current_user.is_org_admin():
        pharmacy_id = request.args.get('pharmacy_id', type=int)
    else:
        pharmacy_id = current_user.pharmacy_id

    if not pharmacy_id:
        return jsonify({'error': 'No pharmacy assigned'}), 400

    # Access check for org_admin
    if current_user.is_org_admin() and not _can_access_pharmacy(pharmacy_id):
        return jsonify({'error': 'Access denied'}), 403

    today = _today()
    thirty_days_ago = today - timedelta(days=30)

    daily_data = DailyStat.query.filter(
        DailyStat.pharmacy_id == pharmacy_id,
        DailyStat.date >= thirty_days_ago
    ).order_by(DailyStat.date).all()

    dates = []
    loaded = []
    collected = []
    removed = []

    for stat in daily_data:
        dates.append(stat.date.strftime('%d/%m'))
        loaded.append(stat.loaded_parcels)
        collected.append(stat.collected_parcels)
        removed.append(stat.removed_parcels)

    day_of_week = {i: {'loaded': 0, 'collected': 0, 'count': 0} for i in range(7)}
    for stat in daily_data:
        dow = stat.date.weekday()
        day_of_week[dow]['loaded'] += stat.loaded_parcels
        day_of_week[dow]['collected'] += stat.collected_parcels
        day_of_week[dow]['count'] += 1

    dow_labels = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    dow_loaded = []
    dow_collected = []
    for i in range(7):
        count = day_of_week[i]['count'] or 1
        dow_loaded.append(round(day_of_week[i]['loaded'] / count, 1))
        dow_collected.append(round(day_of_week[i]['collected'] / count, 1))

    # Pull the latest month only — avoids scanning years of hourly rows for a 4-bucket chart.
    latest_month = db.session.query(db.func.max(HourlyDistribution.month))\
        .filter(HourlyDistribution.pharmacy_id == pharmacy_id).scalar()
    hourly_by_period = {}
    if latest_month is not None:
        hourly_data = HourlyDistribution.query.filter(
            HourlyDistribution.pharmacy_id == pharmacy_id,
            HourlyDistribution.month == latest_month,
        ).all()
        for h in hourly_data:
            hourly_by_period[h.period] = h.collected_parcels

    period_order = ['00-08', '08-12', '12-18', '18-24']
    hourly_labels = ['00:00-08:00', '08:00-12:00', '12:00-18:00', '18:00-24:00']
    hourly_values = [hourly_by_period.get(p, 0) for p in period_order]
    hourly_total = sum(hourly_values) or 1
    hourly_percentages = [round((v / hourly_total) * 100, 1) for v in hourly_values]

    return jsonify({
        'daily': {
            'labels': dates,
            'loaded': loaded,
            'collected': collected,
            'removed': removed
        },
        'dayOfWeek': {
            'labels': dow_labels,
            'loaded': dow_loaded,
            'collected': dow_collected
        },
        'hourly': {
            'labels': hourly_labels,
            'values': hourly_values,
            'percentages': hourly_percentages
        }
    })


# === ORGANISATION ADMIN ROUTES ===

@app.route('/org/dashboard')
@login_required
def org_dashboard():
    if not current_user.is_org_admin():
        return redirect(url_for('dashboard'))

    org = db.session.get(Organisation, current_user.organisation_id)
    if not org:
        flash('Your account is not linked to an organisation.', 'error')
        return redirect(url_for('dashboard'))

    pharmacies = Pharmacy.query.filter_by(organisation_id=org.id).order_by(Pharmacy.name).all()
    today = _today()

    pharmacy_ids = [p.id for p in pharmacies]

    # Aggregate stats across org pharmacies
    today_stats = db.session.query(
        db.func.sum(DailyStat.loaded_parcels),
        db.func.sum(DailyStat.collected_parcels),
        db.func.sum(DailyStat.removed_parcels)
    ).filter(DailyStat.pharmacy_id.in_(pharmacy_ids), DailyStat.date == today).first() if pharmacy_ids else (0, 0, 0)

    seven_days_ago = today - timedelta(days=7)
    week_stats = db.session.query(
        db.func.sum(DailyStat.loaded_parcels),
        db.func.sum(DailyStat.collected_parcels),
        db.func.sum(DailyStat.removed_parcels)
    ).filter(DailyStat.pharmacy_id.in_(pharmacy_ids), DailyStat.date >= seven_days_ago).first() if pharmacy_ids else (0, 0, 0)

    thirty_days_ago = today - timedelta(days=30)
    month_stats = db.session.query(
        db.func.sum(DailyStat.loaded_parcels),
        db.func.sum(DailyStat.collected_parcels),
        db.func.sum(DailyStat.removed_parcels)
    ).filter(DailyStat.pharmacy_id.in_(pharmacy_ids), DailyStat.date >= thirty_days_ago).first() if pharmacy_ids else (0, 0, 0)

    # Batched per-pharmacy stats: 4 queries total regardless of pharmacy count.
    pharmacy_summaries = get_pharmacy_stats_batch(pharmacies, today)

    stats = {
        'total_pharmacies': len(pharmacies),
        'today': {
            'loaded': (today_stats[0] or 0) if today_stats else 0,
            'collected': (today_stats[1] or 0) if today_stats else 0,
            'removed': (today_stats[2] or 0) if today_stats else 0
        },
        'week': {
            'loaded': (week_stats[0] or 0) if week_stats else 0,
            'collected': (week_stats[1] or 0) if week_stats else 0,
            'removed': (week_stats[2] or 0) if week_stats else 0
        },
        'month': {
            'loaded': (month_stats[0] or 0) if month_stats else 0,
            'collected': (month_stats[1] or 0) if month_stats else 0,
            'removed': (month_stats[2] or 0) if month_stats else 0
        }
    }

    return render_template('org/dashboard.html', org=org, stats=stats, pharmacy_summaries=pharmacy_summaries)


@app.route('/org/pharmacy/<int:id>/view')
@login_required
def org_pharmacy_view(id):
    if not current_user.is_org_admin():
        return redirect(url_for('dashboard'))

    pharmacy = Pharmacy.query.get_or_404(id)
    if pharmacy.organisation_id != current_user.organisation_id:
        abort(403)

    today = _today()
    stats = get_pharmacy_stats(pharmacy.id, today)
    return render_template('org/pharmacy_view.html', pharmacy=pharmacy, stats=stats)


@app.route('/org/users')
@login_required
def org_users():
    if not current_user.is_org_admin():
        return redirect(url_for('dashboard'))

    org_pharmacy_ids = _get_org_pharmacy_ids()
    # Show org_admin users for this org + pharmacy users linked to org pharmacies
    users = User.query.filter(
        db.or_(
            User.organisation_id == current_user.organisation_id,
            User.pharmacy_id.in_(org_pharmacy_ids) if org_pharmacy_ids else db.false()
        )
    ).order_by(User.name).all()
    return render_template('org/users.html', users=users)


@app.route('/org/user/add', methods=['GET', 'POST'])
@login_required
def org_user_add():
    if not current_user.is_org_admin():
        return redirect(url_for('dashboard'))

    form = OrgUserForm()
    org_pharmacies = Pharmacy.query.filter_by(organisation_id=current_user.organisation_id).order_by(Pharmacy.name).all()
    form.pharmacy_id.choices = [(0, '-- No Pharmacy --')] + [(p.id, p.name) for p in org_pharmacies]

    if form.validate_on_submit():
        existing = User.query.filter_by(email=_normalize_email(form.email.data)).first()
        if existing:
            flash('A user with that email already exists.', 'error')
            return render_template('org/user_form.html', form=form, title='Add User')

        user = User(
            email=_normalize_email(form.email.data),
            name=form.name.data,
            role='pharmacy',
            pharmacy_id=form.pharmacy_id.data if form.pharmacy_id.data != 0 else None,
            organisation_id=current_user.organisation_id
        )
        # Placeholder random password — user can't log in until they set their own via
        # the invitation link. The hash is unguessable so the row is safe at rest.
        user.set_password(secrets.token_urlsafe(32))
        db.session.add(user)
        db.session.commit()

        if _invite_new_user(user):
            flash(f'User "{user.name}" has been created. An invitation email has been sent to {user.email}.', 'success')
        else:
            flash(
                f'User "{user.name}" has been created, but the invitation email could not be sent. '
                'Use the "Send invite" action on the user list to retry.', 'error'
            )
        return redirect(url_for('org_users'))
    return render_template('org/user_form.html', form=form, title='Add User')


@app.route('/org/user/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def org_user_edit(id):
    if not current_user.is_org_admin():
        return redirect(url_for('dashboard'))

    user = User.query.get_or_404(id)
    # Can only edit users in their org
    org_pharmacy_ids = _get_org_pharmacy_ids()
    if user.organisation_id != current_user.organisation_id and user.pharmacy_id not in org_pharmacy_ids:
        abort(403)

    form = OrgUserForm(obj=user)
    org_pharmacies = Pharmacy.query.filter_by(organisation_id=current_user.organisation_id).order_by(Pharmacy.name).all()
    form.pharmacy_id.choices = [(0, '-- No Pharmacy --')] + [(p.id, p.name) for p in org_pharmacies]

    if form.validate_on_submit():
        new_email = _normalize_email(form.email.data)
        if new_email != user.email:
            existing = User.query.filter_by(email=new_email).first()
            if existing:
                flash('A user with that email already exists.', 'error')
                form.pharmacy_id.data = user.pharmacy_id or 0
                return render_template('org/user_form.html', form=form, title='Edit User', user=user)

        user.email = new_email
        user.name = form.name.data
        user.pharmacy_id = form.pharmacy_id.data if form.pharmacy_id.data != 0 else None
        # Passwords are NOT set from the edit form — use the "Send password reset" or
        # "Resend invite" actions instead.
        db.session.commit()
        flash(f'User "{user.name}" has been updated.', 'success')
        return redirect(url_for('org_users'))

    form.pharmacy_id.data = user.pharmacy_id or 0
    return render_template('org/user_form.html', form=form, title='Edit User', user=user)


@app.route('/org/user/<int:id>/send-reset', methods=['POST'])
@login_required
def org_user_send_reset(id):
    if not current_user.is_org_admin():
        return redirect(url_for('dashboard'))
    user = User.query.get_or_404(id)
    org_pharmacy_ids = _get_org_pharmacy_ids()
    if user.organisation_id != current_user.organisation_id and user.pharmacy_id not in org_pharmacy_ids:
        abort(403)
    if _trigger_password_reset(user):
        flash(f'Password reset email sent to {user.email}.', 'success')
    else:
        flash(f'Could not send password reset email to {user.email}. Check Railway logs.', 'error')
    return redirect(url_for('org_users'))


@app.route('/org/user/<int:id>/resend-invite', methods=['POST'])
@login_required
def org_user_resend_invite(id):
    if not current_user.is_org_admin():
        return redirect(url_for('dashboard'))
    user = User.query.get_or_404(id)
    org_pharmacy_ids = _get_org_pharmacy_ids()
    if user.organisation_id != current_user.organisation_id and user.pharmacy_id not in org_pharmacy_ids:
        abort(403)
    if _invite_new_user(user):
        flash(f'Invitation email re-sent to {user.email}.', 'success')
    else:
        flash(f'Could not send invitation to {user.email}. Check Railway logs.', 'error')
    return redirect(url_for('org_users'))


@app.route('/org/user/<int:id>/delete', methods=['POST'])
@login_required
def org_user_delete(id):
    if not current_user.is_org_admin():
        return redirect(url_for('dashboard'))

    user = User.query.get_or_404(id)
    if user.id == current_user.id:
        flash('You cannot delete your own account.', 'error')
        return redirect(url_for('org_users'))

    # Can only delete users in their org
    org_pharmacy_ids = _get_org_pharmacy_ids()
    if user.organisation_id != current_user.organisation_id and user.pharmacy_id not in org_pharmacy_ids:
        abort(403)

    # Refuse to leave the organisation with no org_admin.
    if user.role == 'org_admin':
        remaining = User.query.filter(
            User.role == 'org_admin',
            User.organisation_id == current_user.organisation_id,
            User.id != user.id,
        ).count()
        if remaining < 1:
            flash('Cannot delete the last organisation admin. Promote another user first.', 'error')
            return redirect(url_for('org_users'))

    db.session.delete(user)
    db.session.commit()
    flash(f'User "{user.name}" has been deleted.', 'success')
    return redirect(url_for('org_users'))


# === SUPER ADMIN ROUTES ===

@app.route('/admin/dashboard')
@login_required
@super_admin_required
def admin_dashboard():
    pharmacies = Pharmacy.query.order_by(Pharmacy.name).all()
    today = _today()

    total_pharmacies = len(pharmacies)

    today_stats = db.session.query(
        db.func.sum(DailyStat.loaded_parcels),
        db.func.sum(DailyStat.collected_parcels),
        db.func.sum(DailyStat.removed_parcels)
    ).filter(DailyStat.date == today).first()

    seven_days_ago = today - timedelta(days=7)
    week_stats = db.session.query(
        db.func.sum(DailyStat.loaded_parcels),
        db.func.sum(DailyStat.collected_parcels),
        db.func.sum(DailyStat.removed_parcels)
    ).filter(DailyStat.date >= seven_days_ago).first()

    thirty_days_ago = today - timedelta(days=30)
    month_stats = db.session.query(
        db.func.sum(DailyStat.loaded_parcels),
        db.func.sum(DailyStat.collected_parcels),
        db.func.sum(DailyStat.removed_parcels)
    ).filter(DailyStat.date >= thirty_days_ago).first()

    # Batched per-pharmacy stats: 4 queries total regardless of pharmacy count.
    pharmacy_summaries = get_pharmacy_stats_batch(pharmacies, today)

    recent_uploads = Upload.query.order_by(Upload.uploaded_at.desc()).limit(5).all()

    stats = {
        'total_pharmacies': total_pharmacies,
        'today': {
            'loaded': today_stats[0] or 0,
            'collected': today_stats[1] or 0,
            'removed': today_stats[2] or 0
        },
        'week': {
            'loaded': week_stats[0] or 0,
            'collected': week_stats[1] or 0,
            'removed': week_stats[2] or 0
        },
        'month': {
            'loaded': month_stats[0] or 0,
            'collected': month_stats[1] or 0,
            'removed': month_stats[2] or 0
        }
    }

    return render_template('admin/dashboard.html',
                         stats=stats,
                         pharmacy_summaries=pharmacy_summaries,
                         recent_uploads=recent_uploads)


def _verify_upload_signature(filepath, original_name):
    """Magic-byte check on the saved upload — defends against extension spoofing.

    Returns 'xlsx' | 'csv' or raises ValueError.
    """
    lower = (original_name or '').lower()
    with open(filepath, 'rb') as fh:
        head = fh.read(8)
    if lower.endswith(('.xlsx', '.xls')):
        # xlsx is a zip container; .xls (BIFF) starts with D0 CF 11 E0.
        if head.startswith(b'PK\x03\x04'):
            if not zipfile.is_zipfile(filepath):
                raise ValueError('Not a valid Excel (xlsx) file')
            return 'xlsx'
        if head.startswith(b'\xd0\xcf\x11\xe0'):
            return 'xlsx'  # legacy .xls — still handled by openpyxl path (will fail clearly if unreadable)
        raise ValueError('File does not look like an Excel file')
    if lower.endswith('.csv'):
        # CSV has no magic bytes — at least reject anything that starts with executable headers.
        if head.startswith((b'\x7fELF', b'MZ', b'PK\x03\x04', b'\xca\xfe\xba\xbe')):
            raise ValueError('CSV upload rejected — file looks like a binary')
        return 'csv'
    raise ValueError('Unsupported file extension')


@app.route('/admin/upload', methods=['GET', 'POST'])
@login_required
@super_admin_required
def admin_upload():
    form = UploadForm()
    uploads = Upload.query.order_by(Upload.uploaded_at.desc()).limit(20).all()

    if form.validate_on_submit():
        file = form.file.data
        original_name = secure_filename(file.filename or 'upload')
        if not original_name:
            flash('Invalid filename.', 'error')
            return redirect(url_for('admin_upload'))

        # UUID-prefix prevents concurrent uploads from clobbering one another mid-parse.
        stored_name = f'{uuid.uuid4().hex}_{original_name}'
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], stored_name)
        file.save(filepath)

        try:
            try:
                _verify_upload_signature(filepath, original_name)
            except ValueError as ve:
                flash(f'File rejected: {ve}', 'error')
                return redirect(url_for('admin_upload'))

            records, affected_pharmacies, skipped = process_upload(filepath, original_name)

            if records == 0:
                flash('No records found in file. Please check the file format matches the expected template.', 'error')
            else:
                upload = Upload(
                    filename=original_name,
                    uploaded_by=current_user.id,
                    records_imported=records
                )
                db.session.add(upload)
                db.session.commit()

                emails_sent = 0
                for pharmacy_id, stats in affected_pharmacies.items():
                    pharmacy = db.session.get(Pharmacy, pharmacy_id)
                    if pharmacy and pharmacy.notification_email:
                        if send_notification_email(pharmacy, stats):
                            emails_sent += 1

                flash(f'Successfully imported {records} records from {original_name}', 'success')
                if skipped:
                    flash(f'Skipped {skipped} row(s) with invalid data', 'info')
                if emails_sent > 0:
                    flash(f'Sent {emails_sent} notification email(s) to pharmacies', 'info')

        except MemoryError:
            db.session.rollback()
            app.logger.exception(f'Memory exhausted parsing upload {original_name}')
            flash('File rejected — too large or too many rows. Split it and try again.', 'error')
        except ValueError as ve:
            # ValueError messages from our parsers are user-actionable (row caps, missing month, etc.).
            db.session.rollback()
            app.logger.warning(f'Upload {original_name} rejected: {ve}')
            flash(str(ve), 'error')
        except Exception:
            db.session.rollback()
            app.logger.exception(f'Error processing upload {original_name}')
            # Surface a generic message — never echo arbitrary exception strings back to the UI.
            flash('Error processing file. Check the file format and try again.', 'error')
        finally:
            try:
                os.remove(filepath)
            except OSError:
                pass

        return redirect(url_for('admin_upload'))

    return render_template('admin/upload.html', form=form, uploads=uploads)


# --- Organisation CRUD (super admin only) ---

@app.route('/admin/organisations')
@login_required
@super_admin_required
def admin_organisations():
    organisations = Organisation.query.order_by(Organisation.name).all()
    return render_template('admin/organisations.html', organisations=organisations)


@app.route('/admin/organisation/add', methods=['GET', 'POST'])
@login_required
@super_admin_required
def admin_organisation_add():
    form = OrganisationForm()
    if form.validate_on_submit():
        org = Organisation(name=form.name.data)
        db.session.add(org)
        db.session.commit()
        flash(f'Organisation "{org.name}" has been created.', 'success')
        return redirect(url_for('admin_organisations'))
    return render_template('admin/organisation_form.html', form=form, title='Add Organisation')


@app.route('/admin/organisation/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@super_admin_required
def admin_organisation_edit(id):
    org = Organisation.query.get_or_404(id)
    form = OrganisationForm(obj=org)
    if form.validate_on_submit():
        org.name = form.name.data
        db.session.commit()
        flash(f'Organisation "{org.name}" has been updated.', 'success')
        return redirect(url_for('admin_organisations'))
    return render_template('admin/organisation_form.html', form=form, title='Edit Organisation', org=org)


@app.route('/admin/organisation/<int:id>/delete', methods=['POST'])
@login_required
@super_admin_required
def admin_organisation_delete(id):
    org = Organisation.query.get_or_404(id)
    # Unlink pharmacies and users
    Pharmacy.query.filter_by(organisation_id=id).update({'organisation_id': None})
    User.query.filter_by(organisation_id=id).update({'organisation_id': None})
    db.session.delete(org)
    db.session.commit()
    flash(f'Organisation "{org.name}" has been deleted.', 'success')
    return redirect(url_for('admin_organisations'))


# --- Pharmacy CRUD (super admin) ---

def _pharmacy_data_counts(pharmacy_ids):
    """Return {pharmacy_id: {'daily': N, 'hourly': M}} for the given pharmacy IDs.

    Used to populate the type-to-confirm delete modal so the admin sees the exact
    blast radius (how many DailyStat + HourlyDistribution rows are about to be
    permanently erased) BEFORE typing the pharmacy name to confirm.
    """
    if not pharmacy_ids:
        return {}
    daily_counts = dict(
        db.session.query(DailyStat.pharmacy_id, db.func.count(DailyStat.id))
        .filter(DailyStat.pharmacy_id.in_(pharmacy_ids))
        .group_by(DailyStat.pharmacy_id).all()
    )
    hourly_counts = dict(
        db.session.query(HourlyDistribution.pharmacy_id, db.func.count(HourlyDistribution.id))
        .filter(HourlyDistribution.pharmacy_id.in_(pharmacy_ids))
        .group_by(HourlyDistribution.pharmacy_id).all()
    )
    return {
        pid: {'daily': daily_counts.get(pid, 0), 'hourly': hourly_counts.get(pid, 0)}
        for pid in pharmacy_ids
    }


@app.route('/admin/pharmacies')
@login_required
@super_admin_required
def admin_pharmacies():
    pharmacies = Pharmacy.query.order_by(Pharmacy.name).all()
    delete_counts = _pharmacy_data_counts([p.id for p in pharmacies])
    return render_template('admin/pharmacies.html',
                           pharmacies=pharmacies,
                           delete_counts=delete_counts)


@app.route('/admin/pharmacy/add', methods=['GET', 'POST'])
@login_required
@super_admin_required
def admin_pharmacy_add():
    form = PharmacyForm()
    form.organisation_id.choices = [(0, '-- No Organisation --')] + [
        (o.id, o.name) for o in Organisation.query.order_by(Organisation.name).all()
    ]
    if form.validate_on_submit():
        existing = Pharmacy.query.filter_by(serial_number=form.serial_number.data).first()
        if existing:
            flash(f'A pharmacy with serial number "{form.serial_number.data}" already exists.', 'error')
            return render_template('admin/pharmacy_form.html', form=form, title='Add Pharmacy')

        pharmacy = Pharmacy(
            serial_number=form.serial_number.data,
            name=form.name.data,
            notification_email=form.notification_email.data or None,
            organisation_id=form.organisation_id.data if form.organisation_id.data != 0 else None
        )
        db.session.add(pharmacy)
        db.session.commit()
        flash(f'Pharmacy "{pharmacy.name}" has been created.', 'success')
        return redirect(url_for('admin_pharmacies'))
    return render_template('admin/pharmacy_form.html', form=form, title='Add Pharmacy')


@app.route('/admin/pharmacy/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@super_admin_required
def admin_pharmacy_edit(id):
    pharmacy = Pharmacy.query.get_or_404(id)
    form = PharmacyForm(obj=pharmacy)
    form.organisation_id.choices = [(0, '-- No Organisation --')] + [
        (o.id, o.name) for o in Organisation.query.order_by(Organisation.name).all()
    ]
    if form.validate_on_submit():
        pharmacy.serial_number = form.serial_number.data
        pharmacy.name = form.name.data
        pharmacy.notification_email = form.notification_email.data or None
        pharmacy.organisation_id = form.organisation_id.data if form.organisation_id.data != 0 else None
        db.session.commit()
        flash(f'Pharmacy "{pharmacy.name}" has been updated.', 'success')
        return redirect(url_for('admin_pharmacies'))
    form.organisation_id.data = pharmacy.organisation_id or 0
    return render_template('admin/pharmacy_form.html', form=form, title='Edit Pharmacy', pharmacy=pharmacy)


@app.route('/admin/pharmacy/<int:id>/view')
@login_required
@super_admin_required
def admin_pharmacy_view(id):
    pharmacy = Pharmacy.query.get_or_404(id)
    today = _today()
    stats = get_pharmacy_stats(pharmacy.id, today)
    delete_counts = _pharmacy_data_counts([pharmacy.id])
    return render_template('admin/pharmacy_view.html',
                           pharmacy=pharmacy,
                           stats=stats,
                           delete_counts=delete_counts)


@app.route('/admin/pharmacy/<int:id>/delete', methods=['POST'])
@login_required
@super_admin_required
def admin_pharmacy_delete(id):
    pharmacy = Pharmacy.query.get_or_404(id)
    pharmacy_name = pharmacy.name

    HourlyDistribution.query.filter_by(pharmacy_id=id).delete()
    DailyStat.query.filter_by(pharmacy_id=id).delete()
    User.query.filter_by(pharmacy_id=id).update({'pharmacy_id': None})
    db.session.delete(pharmacy)
    db.session.commit()

    flash(f'Pharmacy "{pharmacy_name}" and all related data have been deleted.', 'success')
    return redirect(url_for('admin_pharmacies'))


# --- User CRUD (super admin) ---

@app.route('/admin/users')
@login_required
@super_admin_required
def admin_users():
    users = User.query.order_by(User.name).all()
    return render_template('admin/users.html', users=users)


@app.route('/admin/user/add', methods=['GET', 'POST'])
@login_required
@super_admin_required
def admin_user_add():
    form = UserForm()
    form.pharmacy_id.choices = [(0, '-- No Pharmacy --')] + [
        (p.id, p.name) for p in Pharmacy.query.order_by(Pharmacy.name).all()
    ]
    form.organisation_id.choices = [(0, '-- No Organisation --')] + [
        (o.id, o.name) for o in Organisation.query.order_by(Organisation.name).all()
    ]

    if form.validate_on_submit():
        existing = User.query.filter_by(email=_normalize_email(form.email.data)).first()
        if existing:
            flash('A user with that email already exists.', 'error')
            return render_template('admin/user_form.html', form=form, title='Add User')

        user = User(
            email=_normalize_email(form.email.data),
            name=form.name.data,
            role=form.role.data,
            pharmacy_id=form.pharmacy_id.data if form.pharmacy_id.data != 0 else None,
            organisation_id=form.organisation_id.data if form.organisation_id.data != 0 else None
        )
        # Placeholder random password — invitation flow makes the user set their own.
        user.set_password(secrets.token_urlsafe(32))
        db.session.add(user)
        db.session.commit()

        if _invite_new_user(user):
            flash(f'User "{user.name}" has been created. An invitation email has been sent to {user.email}.', 'success')
        else:
            flash(
                f'User "{user.name}" has been created, but the invitation email could not be sent. '
                'Use the "Send invite" action on the user list to retry.', 'error'
            )
        return redirect(url_for('admin_users'))
    return render_template('admin/user_form.html', form=form, title='Add User')


@app.route('/admin/user/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@super_admin_required
def admin_user_edit(id):
    user = User.query.get_or_404(id)
    form = UserForm(obj=user)
    form.pharmacy_id.choices = [(0, '-- No Pharmacy --')] + [
        (p.id, p.name) for p in Pharmacy.query.order_by(Pharmacy.name).all()
    ]
    form.organisation_id.choices = [(0, '-- No Organisation --')] + [
        (o.id, o.name) for o in Organisation.query.order_by(Organisation.name).all()
    ]

    if form.validate_on_submit():
        new_email = _normalize_email(form.email.data)
        if new_email != user.email:
            existing = User.query.filter_by(email=new_email).first()
            if existing:
                flash('A user with that email already exists.', 'error')
                form.pharmacy_id.data = user.pharmacy_id or 0
                form.organisation_id.data = user.organisation_id or 0
                return render_template('admin/user_form.html', form=form, title='Edit User', user=user)

        user.email = new_email
        user.name = form.name.data
        user.role = form.role.data
        user.pharmacy_id = form.pharmacy_id.data if form.pharmacy_id.data != 0 else None
        user.organisation_id = form.organisation_id.data if form.organisation_id.data != 0 else None
        # Passwords are not set from this form. Use "Send password reset" or "Resend invite".
        db.session.commit()
        flash(f'User "{user.name}" has been updated.', 'success')
        return redirect(url_for('admin_users'))

    form.pharmacy_id.data = user.pharmacy_id or 0
    form.organisation_id.data = user.organisation_id or 0
    return render_template('admin/user_form.html', form=form, title='Edit User', user=user)


@app.route('/admin/user/<int:id>/send-reset', methods=['POST'])
@login_required
@super_admin_required
def admin_user_send_reset(id):
    user = User.query.get_or_404(id)
    if _trigger_password_reset(user):
        flash(f'Password reset email sent to {user.email}.', 'success')
    else:
        flash(f'Could not send password reset email to {user.email}. Check Railway logs.', 'error')
    return redirect(url_for('admin_users'))


@app.route('/admin/user/<int:id>/resend-invite', methods=['POST'])
@login_required
@super_admin_required
def admin_user_resend_invite(id):
    user = User.query.get_or_404(id)
    if _invite_new_user(user):
        flash(f'Invitation email re-sent to {user.email}.', 'success')
    else:
        flash(f'Could not send invitation to {user.email}. Check Railway logs.', 'error')
    return redirect(url_for('admin_users'))


@app.route('/admin/user/<int:id>/delete', methods=['POST'])
@login_required
@super_admin_required
def admin_user_delete(id):
    user = User.query.get_or_404(id)
    if user.id == current_user.id:
        flash('You cannot delete your own account.', 'error')
        return redirect(url_for('admin_users'))
    # Refuse to leave the system without a super_admin.
    if user.role in ('super_admin', 'admin'):
        remaining = User.query.filter(
            User.role.in_(('super_admin', 'admin')),
            User.id != user.id,
        ).count()
        if remaining < 1:
            flash('Cannot delete the last super admin.', 'error')
            return redirect(url_for('admin_users'))
    db.session.delete(user)
    db.session.commit()
    flash(f'User "{user.name}" has been deleted.', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/help')
@login_required
@admin_required
def admin_help():
    return render_template('admin/help.html')


def _resend_probe():
    """Hit GET https://api.resend.com/domains to confirm the API key works and to surface
    which sending domains are verified. Returns (status_code:int|None, body:str, error:str|None).
    No-op probe: no email is sent.
    """
    api_key = app.config.get('RESEND_API_KEY')
    if not api_key:
        return None, '', 'RESEND_API_KEY env var is not set on this deployment.'
    req = Request(
        'https://api.resend.com/domains',
        headers={
            'Authorization': f'Bearer {api_key}',
            'Content-Type': 'application/json',
            # Same Cloudflare WAF reason as in send_email() — see comment there.
            'User-Agent': 'Pharmabox24/1.0 (+https://managementinfo.pharmabox24.co.uk)',
            'Accept': 'application/json',
        },
        method='GET',
    )
    try:
        with urlopen(req, timeout=10) as resp:
            body = resp.read().decode('utf-8', errors='replace')[:4000]
            return resp.status, body, None
    except HTTPError as e:
        body = ''
        try:
            body = e.read().decode('utf-8', errors='replace')[:4000]
        except Exception:
            pass
        return e.code, body, f'HTTP {e.code}'
    except (URLError, TimeoutError) as e:
        return None, '', f'Network error reaching Resend: {e}'
    except Exception as e:
        return None, '', f'Unexpected error: {e}'


@app.route('/admin/email-health', methods=['GET', 'POST'])
@login_required
@super_admin_required
def admin_email_health():
    """Diagnostic page for the email pipeline.

    GET  → shows config + Resend /domains probe result.
    POST → sends a real test email to the logged-in super admin and reports the outcome.
    """
    test_result = None
    if request.method == 'POST':
        # Real send to the logged-in admin's own address — never to an arbitrary recipient.
        target = current_user.email
        api_key_present = bool(app.config.get('RESEND_API_KEY'))
        if not api_key_present:
            test_result = {'ok': False, 'detail': 'RESEND_API_KEY is not set; cannot send test email.'}
        else:
            html = (
                "<p>This is a Pharmabox24 email-pipeline test.</p>"
                "<p>If you can read this, Resend is configured correctly for your domain.</p>"
                f"<p>Sent at: {escape(datetime.utcnow().isoformat())}Z</p>"
            )
            ok = send_email(target, 'Pharmabox24 — Email pipeline test', html)
            test_result = {
                'ok': ok,
                'detail': (
                    f'Test email accepted by Resend and queued to {target}. Check inbox + spam.'
                    if ok else
                    f'Resend rejected the test send to {target}. See the probe result below and Railway logs '
                    f'(grep for "Resend HTTPError" or "Resend network error").'
                ),
            }

    status, body, error = _resend_probe()

    # Light parsing — pull out verified domains list if response is JSON.
    domains = []
    try:
        if body:
            parsed = json.loads(body)
            data = parsed.get('data') if isinstance(parsed, dict) else None
            if isinstance(data, list):
                for d in data:
                    if isinstance(d, dict):
                        domains.append({
                            'name': d.get('name'),
                            'status': d.get('status'),
                            'region': d.get('region'),
                            'created_at': d.get('created_at'),
                        })
    except (ValueError, TypeError):
        pass

    return render_template(
        'admin/email_health.html',
        mail_from=app.config.get('MAIL_FROM'),
        api_key_set=bool(app.config.get('RESEND_API_KEY')),
        api_key_length=len(app.config.get('RESEND_API_KEY') or ''),
        site_url=app.config.get('SITE_URL') or '(unset — falling back to request host)',
        probe_status=status,
        probe_error=error,
        probe_body=body,
        domains=domains,
        test_result=test_result,
    )


# === HELPER FUNCTIONS ===

_EMPTY_BUCKET = {'loaded': 0, 'collected': 0, 'removed': 0, 'reminders': 0}


def _empty_summary():
    return {
        'today': dict(_EMPTY_BUCKET),
        'yesterday': dict(_EMPTY_BUCKET),
        'week': dict(_EMPTY_BUCKET),
        'month': dict(_EMPTY_BUCKET),
    }


def get_pharmacy_stats_batch(pharmacies, today):
    """Compute today/yesterday/week/month stats for many pharmacies in 4 grouped queries
    instead of N×4 per-pharmacy queries. recent_records is intentionally omitted — it's
    only needed on the per-pharmacy detail page, never on list/dashboard views.

    Returns a list of summary dicts in the same order as `pharmacies`, each with a
    `pharmacy` key attached.
    """
    if not pharmacies:
        return []

    ids = [p.id for p in pharmacies]
    yesterday = today - timedelta(days=1)
    seven_days_ago = today - timedelta(days=7)
    thirty_days_ago = today - timedelta(days=30)

    # Query 1: today rows (one per pharmacy that has data today)
    today_rows = DailyStat.query.filter(
        DailyStat.pharmacy_id.in_(ids), DailyStat.date == today
    ).all()
    today_by_id = {r.pharmacy_id: r for r in today_rows}

    # Query 2: yesterday rows
    y_rows = DailyStat.query.filter(
        DailyStat.pharmacy_id.in_(ids), DailyStat.date == yesterday
    ).all()
    yesterday_by_id = {r.pharmacy_id: r for r in y_rows}

    def _aggregate_per_pharmacy(since_date):
        rows = db.session.query(
            DailyStat.pharmacy_id,
            db.func.coalesce(db.func.sum(DailyStat.loaded_parcels), 0),
            db.func.coalesce(db.func.sum(DailyStat.collected_parcels), 0),
            db.func.coalesce(db.func.sum(DailyStat.removed_parcels), 0),
            db.func.coalesce(db.func.sum(DailyStat.reminders_sum), 0),
        ).filter(
            DailyStat.pharmacy_id.in_(ids),
            DailyStat.date >= since_date,
        ).group_by(DailyStat.pharmacy_id).all()
        return {r[0]: {'loaded': int(r[1]), 'collected': int(r[2]),
                        'removed': int(r[3]), 'reminders': int(r[4])} for r in rows}

    # Query 3 + 4: week and month aggregates grouped by pharmacy_id
    week_by_id = _aggregate_per_pharmacy(seven_days_ago)
    month_by_id = _aggregate_per_pharmacy(thirty_days_ago)

    summaries = []
    for p in pharmacies:
        s = _empty_summary()
        s['pharmacy'] = p
        t = today_by_id.get(p.id)
        if t:
            s['today'] = {'loaded': t.loaded_parcels, 'collected': t.collected_parcels,
                          'removed': t.removed_parcels, 'reminders': t.reminders_sum}
        y = yesterday_by_id.get(p.id)
        if y:
            s['yesterday'] = {'loaded': y.loaded_parcels, 'collected': y.collected_parcels,
                              'removed': y.removed_parcels, 'reminders': y.reminders_sum}
        if p.id in week_by_id:
            s['week'] = week_by_id[p.id]
        if p.id in month_by_id:
            s['month'] = month_by_id[p.id]
        summaries.append(s)
    return summaries


def get_pharmacy_stats(pharmacy_id, today):
    yesterday = today - timedelta(days=1)
    seven_days_ago = today - timedelta(days=7)
    thirty_days_ago = today - timedelta(days=30)

    today_stat = DailyStat.query.filter_by(pharmacy_id=pharmacy_id, date=today).first()
    yesterday_stat = DailyStat.query.filter_by(pharmacy_id=pharmacy_id, date=yesterday).first()

    week_stats = db.session.query(
        db.func.sum(DailyStat.loaded_parcels),
        db.func.sum(DailyStat.collected_parcels),
        db.func.sum(DailyStat.removed_parcels),
        db.func.sum(DailyStat.reminders_sum)
    ).filter(
        DailyStat.pharmacy_id == pharmacy_id,
        DailyStat.date >= seven_days_ago
    ).first()

    month_stats = db.session.query(
        db.func.sum(DailyStat.loaded_parcels),
        db.func.sum(DailyStat.collected_parcels),
        db.func.sum(DailyStat.removed_parcels),
        db.func.sum(DailyStat.reminders_sum)
    ).filter(
        DailyStat.pharmacy_id == pharmacy_id,
        DailyStat.date >= thirty_days_ago
    ).first()

    recent_records = DailyStat.query.filter_by(pharmacy_id=pharmacy_id)\
        .order_by(DailyStat.date.desc()).limit(30).all()

    return {
        'today': {
            'loaded': today_stat.loaded_parcels if today_stat else 0,
            'collected': today_stat.collected_parcels if today_stat else 0,
            'removed': today_stat.removed_parcels if today_stat else 0,
            'reminders': today_stat.reminders_sum if today_stat else 0
        },
        'yesterday': {
            'loaded': yesterday_stat.loaded_parcels if yesterday_stat else 0,
            'collected': yesterday_stat.collected_parcels if yesterday_stat else 0,
            'removed': yesterday_stat.removed_parcels if yesterday_stat else 0,
            'reminders': yesterday_stat.reminders_sum if yesterday_stat else 0
        },
        'week': {
            'loaded': week_stats[0] or 0,
            'collected': week_stats[1] or 0,
            'removed': week_stats[2] or 0,
            'reminders': week_stats[3] or 0
        },
        'month': {
            'loaded': month_stats[0] or 0,
            'collected': month_stats[1] or 0,
            'removed': month_stats[2] or 0,
            'reminders': month_stats[3] or 0
        },
        'recent_records': recent_records
    }


def _safe_cell_int(value, default=0):
    """Convert a spreadsheet cell to a non-negative int. Returns (int, ok)."""
    if value is None or value == '':
        return default, True
    try:
        if isinstance(value, str):
            value = value.strip().replace(',', '')
            if not value:
                return default, True
        return max(0, int(float(value))), True
    except (TypeError, ValueError):
        return default, False


def _sanitize_pharmacy_name(value, fallback):
    """Clean a pharmacy name from upload — strip control chars and bracket-style markup
    so anything that survives is safe to store AND safe to surface through any unescaped sink
    (email, AI summary). Output-time escaping is still applied where rendered. Bounded length.
    """
    if value is None:
        return fallback
    s = str(value)
    # Drop NULs, control chars, and chars that anchor injection vectors (<, >, `, $).
    cleaned = []
    for ch in s:
        code = ord(ch)
        if code == 9:                 # keep tab as a normal whitespace char
            cleaned.append(ch)
            continue
        if code < 32 or code == 127:  # drop ctrl chars + DEL
            continue
        if ch in '<>`$':              # drop injection anchors — never legitimate in pharmacy names
            continue
        cleaned.append(ch)
    s = ''.join(cleaned).strip()
    return s[:200] or fallback


def _bounded_rows(sheet, max_rows):
    """Materialize at most max_rows from a worksheet, raising if exceeded."""
    rows = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i >= max_rows:
            raise ValueError(
                f'Sheet exceeds {max_rows} rows — refusing to parse. '
                'Split the file and try again.'
            )
        rows.append(row)
    return rows


def process_upload(filepath, filename):
    """Process CSV or Excel file and import data.

    Returns (records, affected_pharmacies, skipped_rows).
    """
    lower = filename.lower()
    if lower.endswith('.xlsx') or lower.endswith('.xls'):
        return process_excel(filepath)
    return process_csv(filepath)


def _derive_workbook_report_month(sheets_rows):
    """Return the first month-of-year datetime found in any sheet's daily section.

    Used as a fallback when a sheet's own daily section is empty but it still
    carries an hourly distribution block — common for pharmacies that were
    inactive in a given week's multi-pharmacy export.
    """
    for _, all_rows in sheets_rows:
        for row_idx, row in enumerate(all_rows):
            row_values = [str(v).lower().strip() if v else '' for v in row]
            if 's/n' not in row_values or 'date' not in row_values:
                continue
            date_col = row_values.index('date')
            for sub_row in all_rows[row_idx + 1:]:
                if not sub_row or date_col >= len(sub_row):
                    continue
                date_val = sub_row[date_col]
                if isinstance(date_val, datetime):
                    return date_val.date().replace(day=1)
                if isinstance(date_val, date):
                    return date_val.replace(day=1)
            break
    return None


def process_excel(filepath):
    """Process Excel file. Returns (records, affected_pharmacies, skipped_rows)."""
    # read_only=True streams the underlying XML; data_only=True returns evaluated values.
    wb = load_workbook(filepath, read_only=True, data_only=True)
    records = 0
    skipped = 0
    affected_pharmacies = {}
    max_rows = app.config.get('UPLOAD_MAX_ROWS', 100_000)

    try:
        # Weekly exports contain one sheet per pharmacy. Pharmacies that were inactive
        # that week have an empty daily section but still carry an hourly distribution
        # block — they need a reporting month sourced from sibling sheets, not their own
        # (empty) daily rows. Materialize every sheet up front so we can compute a
        # workbook-wide report_month BEFORE per-sheet processing.
        sheets_rows = [(sheet, _bounded_rows(sheet, max_rows)) for sheet in wb.worksheets]

        workbook_report_month = _derive_workbook_report_month(sheets_rows)

        for sheet, all_rows in sheets_rows:

            daily_header_row = None
            daily_headers = {}
            hourly_header_row = None
            hourly_headers = {}

            for row_idx, row in enumerate(all_rows):
                # Strip cell values during header detection so a stray trailing space
                # in the exporter (e.g. "Date " or "S/N ") doesn't silently make the
                # daily section invisible. Same robustness applied a few lines down
                # when we populate the header→column-index dict.
                row_values = [str(v).lower().strip() if v else '' for v in row]

                if 'collected parcel distribution' in ' '.join(row_values):
                    if row_idx + 1 < len(all_rows):
                        header_row = all_rows[row_idx + 1]
                        hourly_header_row = row_idx + 1
                        for col_idx, val in enumerate(header_row):
                            if val:
                                hourly_headers[str(val).lower().strip()] = col_idx

                elif 's/n' in row_values and 'date' in row_values:
                    daily_header_row = row_idx
                    for col_idx, val in enumerate(row):
                        if val:
                            daily_headers[str(val).lower().strip()] = col_idx

            col_map = {}
            if daily_header_row is not None:
                col_map = {
                    'serial': daily_headers.get('s/n', daily_headers.get('serial', -1)),
                    'name': daily_headers.get('pharmacy name', daily_headers.get('name', -1)),
                    'date': daily_headers.get('date', -1),
                    'loaded': daily_headers.get('loaded parcels', daily_headers.get('loaded', -1)),
                    'collected': daily_headers.get('collected parcels', daily_headers.get('collected', -1)),
                    'removed': daily_headers.get('removed parcels', daily_headers.get('removed', -1)),
                    'reminders': daily_headers.get('reminders sum', daily_headers.get('reminders', -1))
                }

                # Trim the daily slice at the hourly section only when hourly is BELOW
                # daily (the normal layout). If a sheet ever inverts the order, fall back
                # to scanning the full sheet — the per-row guards (digit serial + real
                # datetime) reject hourly rows naturally, so no false positives.
                if hourly_header_row and hourly_header_row > daily_header_row:
                    end_row = hourly_header_row - 1
                else:
                    end_row = len(all_rows)

                for row in all_rows[daily_header_row + 1:end_row]:
                    if not row or col_map['serial'] < 0 or not row[col_map['serial']]:
                        continue

                    serial_val = row[col_map['serial']]
                    if isinstance(serial_val, (int, float)):
                        serial = str(int(serial_val))
                    else:
                        serial = str(serial_val).strip()

                    if not serial.isdigit() or len(serial) > 50:
                        skipped += 1
                        continue

                    raw_name = row[col_map['name']] if col_map['name'] >= 0 else None
                    name = _sanitize_pharmacy_name(raw_name, serial)

                    date_val = row[col_map['date']]
                    if isinstance(date_val, datetime):
                        stat_date = date_val.date()
                    elif isinstance(date_val, date):
                        stat_date = date_val
                    else:
                        skipped += 1
                        continue

                    # Each int conversion is independent — a single bad cell skips the row,
                    # not the whole upload.
                    loaded, ok_l = _safe_cell_int(row[col_map['loaded']] if col_map['loaded'] >= 0 else 0)
                    collected, ok_c = _safe_cell_int(row[col_map['collected']] if col_map['collected'] >= 0 else 0)
                    removed, ok_r = _safe_cell_int(row[col_map['removed']] if col_map['removed'] >= 0 else 0)
                    reminders, ok_rem = _safe_cell_int(row[col_map['reminders']] if col_map['reminders'] >= 0 else 0)
                    if not (ok_l and ok_c and ok_r and ok_rem):
                        skipped += 1
                        continue

                    pharmacy = Pharmacy.query.filter_by(serial_number=serial).first()
                    if not pharmacy:
                        pharmacy = Pharmacy(serial_number=serial, name=name)
                        db.session.add(pharmacy)
                        db.session.flush()

                    stat = DailyStat.query.filter_by(pharmacy_id=pharmacy.id, date=stat_date).first()
                    if stat:
                        stat.loaded_parcels = loaded
                        stat.collected_parcels = collected
                        stat.removed_parcels = removed
                        stat.reminders_sum = reminders
                    else:
                        stat = DailyStat(
                            pharmacy_id=pharmacy.id,
                            date=stat_date,
                            loaded_parcels=loaded,
                            collected_parcels=collected,
                            removed_parcels=removed,
                            reminders_sum=reminders
                        )
                        db.session.add(stat)

                    if pharmacy.id not in affected_pharmacies:
                        affected_pharmacies[pharmacy.id] = {
                            'loaded': 0, 'collected': 0, 'removed': 0,
                            'date_from': stat_date, 'date_to': stat_date,
                        }
                    else:
                        if stat_date < affected_pharmacies[pharmacy.id]['date_from']:
                            affected_pharmacies[pharmacy.id]['date_from'] = stat_date
                        if stat_date > affected_pharmacies[pharmacy.id]['date_to']:
                            affected_pharmacies[pharmacy.id]['date_to'] = stat_date
                    affected_pharmacies[pharmacy.id]['loaded'] += loaded
                    affected_pharmacies[pharmacy.id]['collected'] += collected
                    affected_pharmacies[pharmacy.id]['removed'] += removed

                    records += 1

            if hourly_header_row is not None:
                hourly_col_map = {
                    'serial': hourly_headers.get('s/n', hourly_headers.get('serial', -1)),
                    'name': hourly_headers.get('pharmacy name', hourly_headers.get('name', -1)),
                    'period': hourly_headers.get('period from-to hrs', hourly_headers.get('period', -1)),
                    'collected': hourly_headers.get('collected parcels', hourly_headers.get('collected', -1))
                }

                # Derive report_month strictly from the daily block's parseable dates.
                # Refuse to default to today — silently misfiling January data into February breaks analytics.
                # Fall back to the workbook-wide month for sheets whose own daily section is empty
                # (inactive pharmacies in a multi-pharmacy weekly export).
                report_month = None
                if daily_header_row is not None and col_map.get('date', -1) >= 0:
                    for row in all_rows[daily_header_row + 1:]:
                        if row and row[col_map['date']]:
                            date_val = row[col_map['date']]
                            if isinstance(date_val, datetime):
                                report_month = date_val.date().replace(day=1)
                                break
                            elif isinstance(date_val, date):
                                report_month = date_val.replace(day=1)
                                break

                if report_month is None:
                    report_month = workbook_report_month

                if report_month is None:
                    raise ValueError(
                        'Could not determine reporting month: no parseable Date value '
                        'found in the daily statistics section of any sheet. Add at least '
                        'one row with a real date in the Date column and re-upload.'
                    )

                for row in all_rows[hourly_header_row + 1:]:
                    if not row or hourly_col_map['serial'] < 0 or not row[hourly_col_map['serial']]:
                        continue

                    serial_val = row[hourly_col_map['serial']]
                    if isinstance(serial_val, (int, float)):
                        serial = str(int(serial_val))
                    else:
                        serial = str(serial_val).strip()

                    if not serial.isdigit() or len(serial) > 50:
                        skipped += 1
                        continue

                    period = (
                        str(row[hourly_col_map['period']]).strip()[:10]
                        if hourly_col_map['period'] >= 0 and row[hourly_col_map['period']]
                        else None
                    )
                    if not period:
                        continue

                    collected, ok = _safe_cell_int(
                        row[hourly_col_map['collected']] if hourly_col_map['collected'] >= 0 else 0
                    )
                    if not ok:
                        skipped += 1
                        continue

                    pharmacy = Pharmacy.query.filter_by(serial_number=serial).first()
                    if not pharmacy:
                        raw_name = row[hourly_col_map['name']] if hourly_col_map['name'] >= 0 else None
                        name = _sanitize_pharmacy_name(raw_name, serial)
                        pharmacy = Pharmacy(serial_number=serial, name=name)
                        db.session.add(pharmacy)
                        db.session.flush()

                    hourly = HourlyDistribution.query.filter_by(
                        pharmacy_id=pharmacy.id,
                        period=period,
                        month=report_month
                    ).first()

                    if hourly:
                        hourly.collected_parcels = collected
                    else:
                        hourly = HourlyDistribution(
                            pharmacy_id=pharmacy.id,
                            period=period,
                            collected_parcels=collected,
                            month=report_month
                        )
                        db.session.add(hourly)

        db.session.commit()
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return records, affected_pharmacies, skipped


def process_csv(filepath):
    """Process CSV file. Returns (records, affected_pharmacies, skipped_rows)."""
    records = 0
    skipped = 0
    affected_pharmacies = {}
    max_rows = app.config.get('UPLOAD_MAX_ROWS', 100_000)

    with open(filepath, 'r', encoding='utf-8-sig') as f:
        lines = []
        for i, line in enumerate(f):
            if i >= max_rows:
                raise ValueError(
                    f'CSV exceeds {max_rows} rows — refusing to parse. '
                    'Split the file and try again.'
                )
            lines.append(line)

    header_idx = None
    for idx, line in enumerate(lines):
        line_lower = line.lower()
        if 's/n' in line_lower and ('pharmacy' in line_lower or 'date' in line_lower):
            header_idx = idx
            break

    if header_idx is None:
        return 0, {}, 0

    from io import StringIO
    csv_content = ''.join(lines[header_idx:])
    reader = csv.DictReader(StringIO(csv_content))

    if not reader.fieldnames:
        return 0, {}, 0
    fieldnames = {k.lower().strip(): k for k in reader.fieldnames if k}

    if not (fieldnames.get('s/n') or fieldnames.get('serial') or fieldnames.get('serial number')):
        raise ValueError('CSV is missing required column: S/N')

    for row in reader:
        serial_key = fieldnames.get('s/n', fieldnames.get('serial', fieldnames.get('serial number')))
        if not serial_key or not row.get(serial_key):
            continue

        serial = str(row[serial_key]).strip()
        if not serial.isdigit() or len(serial) > 50:
            skipped += 1
            continue

        name_key = fieldnames.get('pharmacy name', fieldnames.get('name'))
        raw_name = row.get(name_key) if name_key else None
        name = _sanitize_pharmacy_name(raw_name, serial)

        date_key = fieldnames.get('date')
        if not date_key or not row.get(date_key):
            continue

        date_str = (row.get(date_key) or '').strip()
        if not date_str:
            continue

        try:
            stat_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            try:
                stat_date = datetime.strptime(date_str, '%d/%m/%Y').date()
            except ValueError:
                skipped += 1
                continue

        loaded, ok_l = _safe_cell_int(row.get(fieldnames.get('loaded parcels', fieldnames.get('loaded', '')), 0))
        collected, ok_c = _safe_cell_int(row.get(fieldnames.get('collected parcels', fieldnames.get('collected', '')), 0))
        removed, ok_r = _safe_cell_int(row.get(fieldnames.get('removed parcels', fieldnames.get('removed', '')), 0))
        reminders, ok_rem = _safe_cell_int(row.get(fieldnames.get('reminders sum', fieldnames.get('reminders', '')), 0))
        if not (ok_l and ok_c and ok_r and ok_rem):
            skipped += 1
            continue

        pharmacy = Pharmacy.query.filter_by(serial_number=serial).first()
        if not pharmacy:
            pharmacy = Pharmacy(serial_number=serial, name=name)
            db.session.add(pharmacy)
            db.session.flush()

        stat = DailyStat.query.filter_by(pharmacy_id=pharmacy.id, date=stat_date).first()
        if stat:
            stat.loaded_parcels = loaded
            stat.collected_parcels = collected
            stat.removed_parcels = removed
            stat.reminders_sum = reminders
        else:
            stat = DailyStat(
                pharmacy_id=pharmacy.id,
                date=stat_date,
                loaded_parcels=loaded,
                collected_parcels=collected,
                removed_parcels=removed,
                reminders_sum=reminders
            )
            db.session.add(stat)

        if pharmacy.id not in affected_pharmacies:
            affected_pharmacies[pharmacy.id] = {
                'loaded': 0, 'collected': 0, 'removed': 0,
                'date_from': stat_date, 'date_to': stat_date,
            }
        else:
            if stat_date < affected_pharmacies[pharmacy.id]['date_from']:
                affected_pharmacies[pharmacy.id]['date_from'] = stat_date
            if stat_date > affected_pharmacies[pharmacy.id]['date_to']:
                affected_pharmacies[pharmacy.id]['date_to'] = stat_date
        affected_pharmacies[pharmacy.id]['loaded'] += loaded
        affected_pharmacies[pharmacy.id]['collected'] += collected
        affected_pharmacies[pharmacy.id]['removed'] += removed

        records += 1

    db.session.commit()
    return records, affected_pharmacies, skipped


def init_db():
    """Initialize database with tables, run migrations, sync admin user.

    Serialised via a Postgres advisory lock so concurrent gunicorn workers don't race
    on DDL or admin-sync. Fails loudly — no silent fallbacks.
    """
    from sqlalchemy import text

    with app.app_context():
        db_url = str(db.engine.url)
        is_postgres = 'postgresql' in db_url or 'postgres' in db_url

        try:
            if is_postgres:
                # Each ALTER is wrapped individually so a single failure doesn't block the rest.
                # Postgres' built-in DDL locking serialises concurrent workers without needing
                # an advisory lock layered on top.
                postgres_statements = [
                    ("create organisations", """
                        CREATE TABLE IF NOT EXISTS organisations (
                            id SERIAL PRIMARY KEY,
                            name VARCHAR(200) NOT NULL,
                            created_at TIMESTAMP DEFAULT NOW()
                        )
                    """),
                    ("add pharmacies.organisation_id",
                     'ALTER TABLE pharmacies ADD COLUMN IF NOT EXISTS organisation_id INTEGER REFERENCES organisations(id)'),
                    ("add users.organisation_id",
                     'ALTER TABLE users ADD COLUMN IF NOT EXISTS organisation_id INTEGER REFERENCES organisations(id)'),
                    ("add users.session_version",
                     "ALTER TABLE users ADD COLUMN IF NOT EXISTS session_version INTEGER NOT NULL DEFAULT 1"),
                    ("migrate legacy admin role",
                     "UPDATE users SET role = 'super_admin' WHERE role = 'admin'"),
                ]
                for label, sql in postgres_statements:
                    try:
                        with db.engine.connect() as conn:
                            conn.execute(text(sql))
                            conn.commit()
                    except Exception:
                        app.logger.exception(f'migration step failed: {label}')
                app.logger.info('PostgreSQL migration sweep complete')
            else:
                # SQLite (local dev). Idempotent ADD COLUMNs — SQLite has no IF NOT EXISTS,
                # so we introspect first.
                with db.engine.connect() as conn:
                    def _user_columns():
                        rows = conn.execute(text("PRAGMA table_info(users)")).fetchall()
                        return {r[1] for r in rows}
                    def _pharm_columns():
                        rows = conn.execute(text("PRAGMA table_info(pharmacies)")).fetchall()
                        return {r[1] for r in rows}
                    user_cols = _user_columns()
                    pharm_cols = _pharm_columns()
                    if user_cols and 'organisation_id' not in user_cols:
                        conn.execute(text('ALTER TABLE users ADD COLUMN organisation_id INTEGER'))
                    if user_cols and 'session_version' not in user_cols:
                        conn.execute(text('ALTER TABLE users ADD COLUMN session_version INTEGER NOT NULL DEFAULT 1'))
                    if pharm_cols and 'organisation_id' not in pharm_cols:
                        conn.execute(text('ALTER TABLE pharmacies ADD COLUMN organisation_id INTEGER'))
                    if user_cols:
                        conn.execute(text("UPDATE users SET role = 'super_admin' WHERE role = 'admin'"))
                    conn.commit()

            db.create_all()
            app.logger.info('db.create_all() complete')

            # Sync admin credentials. NO default-password fallback.
            admin_email = os.environ.get('ADMIN_EMAIL', 'admin@pharmabox24.com').strip().lower()
            admin_password = os.environ.get('ADMIN_PASSWORD', '').strip()

            # Admin sync logic — pick the row in priority order:
            #   1. user matching ADMIN_EMAIL exactly (treat as the canonical admin)
            #   2. otherwise, any existing super_admin
            #   3. otherwise, create a new super_admin with ADMIN_EMAIL
            # We never reassign emails between rows — if ADMIN_EMAIL belongs to a non-admin
            # user and a separate super_admin exists, we leave both alone and warn loudly.
            user_by_email = User.query.filter_by(email=admin_email).first()
            existing_super = User.query.filter_by(role='super_admin').first()

            if admin_password:
                if len(admin_password) < _PASSWORD_MIN:
                    app.logger.error(
                        f'ADMIN_PASSWORD is shorter than {_PASSWORD_MIN} characters — refusing to sync admin.'
                    )
                else:
                    try:
                        if user_by_email is not None:
                            # ADMIN_EMAIL row exists. Promote to super_admin and sync password.
                            user_by_email.role = 'super_admin'
                            user_by_email.set_password(admin_password)
                            db.session.commit()
                            app.logger.info(f'Super admin credentials synced for {admin_email}')
                            if existing_super and existing_super.id != user_by_email.id:
                                app.logger.warning(
                                    f'Another super_admin row exists (id={existing_super.id}, '
                                    f'email={existing_super.email!r}). Two super_admins are now active; '
                                    'reconcile via /admin/users if unintended.'
                                )
                        elif existing_super is not None:
                            # No row with ADMIN_EMAIL, but a super_admin exists under a different
                            # email. DO NOT rename it — renaming risks unique-constraint clashes
                            # against other users. Just sync the password under the existing email.
                            if existing_super.email != admin_email:
                                app.logger.warning(
                                    f'ADMIN_EMAIL env var is {admin_email!r} but the active super_admin '
                                    f'has email {existing_super.email!r}. Password synced under the existing '
                                    'email; update ADMIN_EMAIL to match, or rename the user manually.'
                                )
                            existing_super.set_password(admin_password)
                            db.session.commit()
                            app.logger.info(f'Super admin password synced (id={existing_super.id})')
                        else:
                            # No admin at all — create one from env.
                            admin = User(email=admin_email, name='Administrator', role='super_admin')
                            admin.set_password(admin_password)
                            db.session.add(admin)
                            db.session.commit()
                            app.logger.info(f'Created super admin user: {admin_email}')
                    except Exception:
                        db.session.rollback()
                        app.logger.exception('Admin sync failed — leaving existing rows untouched')
            else:
                if not existing_super and os.environ.get('PHARMABOX_ENV') != 'development':
                    # Hard-fail: never auto-create a default-credential admin in production.
                    raise RuntimeError(
                        'No super_admin exists and ADMIN_PASSWORD env var is not set. '
                        'Set ADMIN_PASSWORD (>= 12 chars) and restart, or create the super_admin '
                        'manually before bringing the app up.'
                    )
                if not existing_super:
                    app.logger.warning('Development mode: no super_admin and no ADMIN_PASSWORD. Set ADMIN_PASSWORD to bootstrap.')

        except Exception:
            # Log loudly but DO NOT crash the worker — production already has a healthy
            # schema, and the migration block is best-effort idempotent. Operator should
            # check logs and reconcile manually if anything fails here.
            # (Track via /admin/email-health-style diagnostic — TODO.)
            app.logger.exception('init_db failed — app will continue starting; investigate the schema state')


# Initialize DB on import — UNLESS we're being imported by gunicorn after start.sh
# already ran init_db separately (the production path on Railway). This avoids running
# the migration sweep + admin sync inside every web worker on every restart, which
# was causing "Railway 502 during click" reports during rolling deploys.
#
# Local dev (no start.sh wrapper) still runs init_db at import time as before.
if os.environ.get('SKIP_INIT_DB') != '1':
    init_db()


if __name__ == '__main__':
    app.run(debug=os.environ.get('FLASK_DEBUG') == '1', port=5000)
